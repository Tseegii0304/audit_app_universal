"""
АУДИТЫН ХОУ ПРОТОТИП v3.4
TB + Ledger + Part1 → Бүрэн шинжилгээ
pip install streamlit pandas numpy scikit-learn plotly openpyxl
streamlit run audit_app.py
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from sklearn.ensemble import IsolationForest, RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import StratifiedKFold, cross_val_predict
from sklearn.metrics import precision_score, recall_score, f1_score, roc_auc_score, roc_curve
import warnings, io, re, gzip
from datetime import datetime
from collections import Counter
warnings.filterwarnings('ignore')
try:
    from tab_descriptions import TabDescriptions
except Exception:
    class TabDescriptions:
        def __getattr__(self, name):
            def _noop(*args, **kwargs):
                return None
            return _noop
td = TabDescriptions()
st.set_page_config(page_title="Аудитын ХОУ v3.4", page_icon="🔍", layout="wide")
st.markdown('<h1 style="text-align:center;color:#1565c0">🔍 Аудитын ХОУ Прототип v3.4</h1>', unsafe_allow_html=True)
st.markdown('<p style="text-align:center;color:#666">TB + Ledger + Part1 → Бүрэн шинжилгээ</p>', unsafe_allow_html=True)

with st.sidebar:
    st.header("📌 Цэс")
    page = st.radio("Алхам:", ["1️⃣ Өгөгдөл бэлтгэх", "2️⃣ Шинжилгээ"])

ACCT_RE_B = re.compile(r'Данс:\s*\[([^\]]+)\]\s*(.*)')
ACCT_RE_P = re.compile(r'Данс:\s*(\d{3}-\d{2}-\d{2}-\d{3})\s+(.*)')

def parse_account(text):
    m = ACCT_RE_B.match(text)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    m = ACCT_RE_P.match(text)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return None, None

def safe_float(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0

def process_raw_tb(file_obj):
    import openpyxl
    wb = openpyxl.load_workbook(file_obj, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        try:
            int(float(row[0]))
        except Exception:
            continue
        code = str(row[1]).strip() if row[1] else ''
        if not code or not re.match(r'\d{3}-', code):
            continue
        rows.append({
            'account_code': code,
            'account_name': str(row[2]).strip() if row[2] else '',
            'opening_debit': safe_float(row[3]),
            'opening_credit': safe_float(row[4]),
            'turnover_debit': safe_float(row[5]),
            'turnover_credit': safe_float(row[6]),
            'closing_debit': safe_float(row[7]),
            'closing_credit': safe_float(row[8]),
        })
    wb.close()
    df = pd.DataFrame(rows)
    df['opening_balance_signed'] = df['opening_debit'] - df['opening_credit']
    df['turnover_net_signed'] = df['turnover_debit'] - df['turnover_credit']
    df['closing_balance_signed'] = df['closing_debit'] - df['closing_credit']
    df['net_change_signed'] = df['closing_balance_signed'] - df['opening_balance_signed']
    tb_sum = df[['account_code','account_name','opening_debit','opening_credit','opening_balance_signed',
                  'turnover_debit','turnover_credit','turnover_net_signed',
                  'closing_debit','closing_credit','closing_balance_signed','net_change_signed']].copy()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df[['account_code','account_name','opening_debit','opening_credit','turnover_debit','turnover_credit','closing_debit','closing_credit']].to_excel(w, sheet_name='01_TB_CLEAN', index=False)
        tb_sum.to_excel(w, sheet_name='02_ACCOUNT_SUMMARY', index=False)
    buf.seek(0)
    return buf, tb_sum

COL_PATTERNS = {
    'account_code': ['дансны код','данс код','account code','account no','account number','acc code','код','дебет данс','кредит данс'],
    'account_name': ['дансны нэр','данс нэр','account name','acc name','нэр'],
    'transaction_date': ['огноо','date','transaction date','txn date'],
    'debit_mnt': ['дебит','debit','dt','дт','debit amount'],
    'credit_mnt': ['кредит','credit','ct','кт','credit amount'],
    'amount_mnt': ['мөнгөн дүн','дүн','amount','amt','transaction amount'],
    'balance_mnt': ['үлдэгдэл','balance','bal','ending balance'],
    'counterparty_name': ['харилцагч','counterparty','partner','vendor','customer','байгууллагын нэр'],
    'transaction_description': ['тайлбар','гүйлгээний утга','утга','description','memo','narration'],
    'journal_no': ['журнал','journal','journal no','журналын төрөл'],
    'document_no': ['баримт','document','doc no','баримт №','баримт дугаар'],
    'asset_expense': ['биет ба биет бус хөрөнгийн зардал'],
}
def _match_col(h, field):
    h2 = str(h).lower().strip()
    return any(p in h2 for p in COL_PATTERNS.get(field, []))
def _auto_map(headers):
    m, used = {}, set()
    for f in ['account_code','debit_mnt','credit_mnt','transaction_date','account_name','counterparty_name','transaction_description','balance_mnt','journal_no','document_no']:
        for i, h in enumerate(headers):
            if i in used: continue
            if _match_col(h, f): m[f]=i; used.add(i); break
    return m


def _find_header_row(all_rows, max_scan=30):
    best_i, best_s = 0, 0
    for i, row in enumerate(all_rows[:max_scan]):
        vals = [str(c).strip().lower() for c in row if c is not None]
        score = 0
        for v in vals:
            if 'огноо' in v or 'date' in v: score += 1
            if 'дебет' in v or 'debit' in v: score += 1
            if 'кредит' in v or 'credit' in v: score += 1
            if 'мөнгөн дүн' in v or 'amount' in v: score += 1
            if 'гүйлгээний утга' in v or 'description' in v: score += 1
        if score > best_s:
            best_s, best_i = score, i
    return best_i, best_s

def _build_dual_entry_from_table(file_obj, report_year):
    import openpyxl
    EDT_COLUMNS = ['report_year','account_code','account_name','transaction_no','transaction_date',
                   'journal_no','document_no','counterparty_name','counterparty_id',
                   'transaction_description','debit_mnt','credit_mnt','balance_mnt','month']
    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        all_rows.append(list(row))
        if i >= 5000:
            break
    wb.close()
    if not all_rows:
        return pd.DataFrame(columns=EDT_COLUMNS), 0
    hdr_i, hdr_score = _find_header_row(all_rows)
    if hdr_score < 3:
        return pd.DataFrame(columns=EDT_COLUMNS), 0
    headers = [str(c).strip() if c is not None else f'col_{j}' for j, c in enumerate(all_rows[hdr_i])]
    cm = _auto_map(headers)

    # This specific format has Debit account, Credit account and one Amount column.
    debit_idx = cm.get('debit_mnt')
    credit_idx = cm.get('credit_mnt')
    amount_idx = cm.get('amount_mnt')
    date_idx = cm.get('transaction_date')
    doc_idx = cm.get('document_no')
    cp_idx = cm.get('counterparty_name')
    desc_idx = cm.get('transaction_description')
    j_idx = cm.get('journal_no')

    if debit_idx is None or credit_idx is None or amount_idx is None:
        return pd.DataFrame(columns=EDT_COLUMNS), 0

    rows_out = []
    tx_counter = 0
    for row in all_rows[hdr_i+1:]:
        if not row or all(c is None or str(c).strip()=='' for c in row):
            continue
        debit_acct = str(row[debit_idx]).strip() if debit_idx < len(row) and row[debit_idx] is not None else ''
        credit_acct = str(row[credit_idx]).strip() if credit_idx < len(row) and row[credit_idx] is not None else ''
        amount = safe_float(row[amount_idx]) if amount_idx < len(row) else 0.0
        if not debit_acct or not credit_acct or amount == 0:
            continue

        raw_date = row[date_idx] if date_idx is not None and date_idx < len(row) else ''
        if isinstance(raw_date, datetime):
            tx_date = raw_date.strftime('%Y-%m-%d')
        else:
            s = str(raw_date).strip() if raw_date is not None else ''
            tx_date = s
            # convert 25.01.02 -> 2025-01-02
            m = re.match(r'^(\d{2})[./-](\d{2})[./-](\d{2})$', s)
            if m:
                yy, mm, dd = m.groups()
                tx_date = f'20{yy}-{mm}-{dd}'

        doc_no = str(row[doc_idx]).strip() if doc_idx is not None and doc_idx < len(row) and row[doc_idx] is not None else ''
        cp_name = str(row[cp_idx]).strip() if cp_idx is not None and cp_idx < len(row) and row[cp_idx] is not None else ''
        desc = str(row[desc_idx]).strip() if desc_idx is not None and desc_idx < len(row) and row[desc_idx] is not None else ''
        journal_no = str(row[j_idx]).strip() if j_idx is not None and j_idx < len(row) and row[j_idx] is not None else ''

        tx_counter += 1
        common = {
            'report_year': str(report_year),
            'transaction_no': str(tx_counter),
            'transaction_date': tx_date,
            'journal_no': journal_no,
            'document_no': doc_no,
            'counterparty_name': cp_name,
            'counterparty_id': '',
            'transaction_description': desc,
            'balance_mnt': 0.0,
            'month': tx_date[:7] if len(tx_date) >= 7 else ''
        }
        rows_out.append({**common, 'account_code': debit_acct, 'account_name': '', 'debit_mnt': amount, 'credit_mnt': 0.0})
        rows_out.append({**common, 'account_code': credit_acct, 'account_name': '', 'debit_mnt': 0.0, 'credit_mnt': amount})

    if rows_out:
        return pd.DataFrame(rows_out)[EDT_COLUMNS], len(rows_out)
    return pd.DataFrame(columns=EDT_COLUMNS), 0

def process_edt(file_obj, report_year):
    """Ямар ч форматын ЕДТ/Ерөнхий журнал уншина."""
    import openpyxl
    EDT_COLUMNS = ['report_year','account_code','account_name','transaction_no','transaction_date',
                   'journal_no','document_no','counterparty_name','counterparty_id',
                   'transaction_description','debit_mnt','credit_mnt','balance_mnt','month']

    # 0-р оролдлого: хүснэгтэн журнал (Дебет/Кредит/Мөнгөн дүн)
    file_obj.seek(0)
    df_journal, cnt_journal = _build_dual_entry_from_table(file_obj, report_year)
    if cnt_journal > 0:
        return df_journal, cnt_journal

    # 1-р оролдлого: Стандарт ЕДТ формат (Данс: [...])
    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_out, cur_code, cur_name = [], None, None
    for row in ws.iter_rows(values_only=True):
        c0 = row[0]
        if c0 is None: continue
        s = str(c0).strip()
        if s.startswith('Данс:'):
            code, name = parse_account(s)
            if code: cur_code, cur_name = code, name
            continue
        if any(s.startswith(x) for x in ['Компани:','ЕРӨНХИЙ','Тайлант','Үүсгэсэн','Журнал:','№','Эцсийн','Дт -','Нийт','Эхний','Нээгээд']) or s in ('Валютаар','Төгрөгөөр',''): continue
        try: tx_no = int(float(c0))
        except: continue
        if cur_code is None: continue
        td = row[1] if len(row)>1 else ''
        tx_date = td.strftime('%Y-%m-%d') if isinstance(td, datetime) else (str(td).strip() if td else '')
        rows_out.append({'report_year':str(report_year),'account_code':cur_code,'account_name':cur_name,
            'transaction_no':str(tx_no),'transaction_date':tx_date,
            'journal_no':str(row[5]).strip() if len(row)>5 and row[5] else '',
            'document_no':str(row[6]).strip() if len(row)>6 and row[6] else '',
            'counterparty_name':str(row[3]).strip() if len(row)>3 and row[3] else '',
            'counterparty_id':str(row[4]).strip() if len(row)>4 and row[4] else '',
            'transaction_description':str(row[7]).strip() if len(row)>7 and row[7] else '',
            'debit_mnt':safe_float(row[9]) if len(row)>9 else 0.0,
            'credit_mnt':safe_float(row[11]) if len(row)>11 else 0.0,
            'balance_mnt':safe_float(row[13]) if len(row)>13 else 0.0,
            'month':tx_date[:7] if len(tx_date)>=7 else ''})
    wb.close()
    if rows_out:
        return pd.DataFrame(rows_out), len(rows_out)

    # 2-р оролдлого: Хүснэгт формат (баганы гарчигтай)
    file_obj.seek(0)
    try:
        raw = file_obj.read(); file_obj.seek(0)
        wb2 = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        ws2 = wb2[wb2.sheetnames[0]]
        allr = []
        for i, row in enumerate(ws2.iter_rows(values_only=True)):
            allr.append(list(row))
            if i >= 500: break
        wb2.close()
        # Гарчиг олох
        best_i, best_s = 0, 0
        for i, row in enumerate(allr[:20]):
            sc = sum(1 for cell in row if cell and any(any(p in str(cell).lower() for p in pats) for pats in COL_PATTERNS.values()))
            if sc > best_s: best_s=sc; best_i=i
        if best_s >= 2:
            headers = [str(c).strip() if c else f'col_{j}' for j, c in enumerate(allr[best_i])]
            cm = _auto_map(headers)
            if 'debit_mnt' in cm or 'credit_mnt' in cm:
                def _gv(row, cm, f, d=''):
                    idx=cm.get(f)
                    if idx is None or idx>=len(row) or row[idx] is None: return d
                    return str(row[idx]).strip()
                rows2 = []
                for row in allr[best_i+1:]:
                    if all(c is None for c in row): continue
                    ac = _gv(row, cm, 'account_code')
                    if not ac or ac in ('None','nan',''): continue
                    db = safe_float(row[cm['debit_mnt']]) if 'debit_mnt' in cm and cm['debit_mnt']<len(row) else 0.0
                    cr = safe_float(row[cm['credit_mnt']]) if 'credit_mnt' in cm and cm['credit_mnt']<len(row) else 0.0
                    if db==0 and cr==0: continue
                    tdi = cm.get('transaction_date')
                    tx_date = ''
                    if tdi is not None and tdi<len(row):
                        td2 = row[tdi]
                        tx_date = td2.strftime('%Y-%m-%d') if isinstance(td2, datetime) else (str(td2).strip()[:10] if td2 else '')
                    rows2.append({'report_year':str(report_year),'account_code':ac,'account_name':_gv(row,cm,'account_name'),
                        'transaction_no':str(len(rows2)+1),'transaction_date':tx_date,
                        'journal_no':_gv(row,cm,'journal_no'),'document_no':_gv(row,cm,'document_no'),
                        'counterparty_name':_gv(row,cm,'counterparty_name'),'counterparty_id':'',
                        'transaction_description':_gv(row,cm,'transaction_description'),
                        'debit_mnt':db,'credit_mnt':cr,
                        'balance_mnt':safe_float(row[cm['balance_mnt']]) if 'balance_mnt' in cm and cm['balance_mnt']<len(row) else 0.0,
                        'month':tx_date[:7] if len(tx_date)>=7 else ''})
                if rows2: return pd.DataFrame(rows2), len(rows2)
    except: pass

    # 3-р оролдлого: pandas-аар шууд
    file_obj.seek(0)
    try:
        df = pd.read_excel(file_obj)
        cm = _auto_map(df.columns.tolist())
        if 'debit_mnt' in cm or 'credit_mnt' in cm:
            rn = {df.columns[idx]: field for field, idx in cm.items()}
            df = df.rename(columns=rn)
            df['report_year'] = str(report_year)
            for c in EDT_COLUMNS:
                if c not in df.columns: df[c] = '' if c in ('account_code','account_name','transaction_description','counterparty_name') else 0
            df['debit_mnt'] = pd.to_numeric(df.get('debit_mnt',0), errors='coerce').fillna(0)
            df['credit_mnt'] = pd.to_numeric(df.get('credit_mnt',0), errors='coerce').fillna(0)
            df = df[(df['debit_mnt']!=0)|(df['credit_mnt']!=0)]
            df['month'] = df['transaction_date'].astype(str).str[:7] if 'transaction_date' in df.columns else ''
            if len(df) > 0: return df[EDT_COLUMNS], len(df)
    except: pass

    return pd.DataFrame(columns=EDT_COLUMNS), 0

def generate_part1(df_led, year):
    df = df_led.copy()
    yr = str(year)
    df['debit_mnt'] = pd.to_numeric(df['debit_mnt'], errors='coerce').fillna(0)
    df['credit_mnt'] = pd.to_numeric(df['credit_mnt'], errors='coerce').fillna(0)
    df['balance_mnt'] = pd.to_numeric(df['balance_mnt'], errors='coerce').fillna(0)
    monthly = df.groupby(['month', 'account_code']).agg(
        total_debit_mnt=('debit_mnt', 'sum'),
        total_credit_mnt=('credit_mnt', 'sum'),
        ending_balance_mnt=('balance_mnt', 'last'),
        transaction_count=('debit_mnt', 'count')
    ).reset_index()
    monthly.insert(0, 'report_year', yr)
    anames = df.groupby('account_code')['account_name'].first()
    acct = df.groupby('account_code').agg(
        total_debit_mnt=('debit_mnt', 'sum'),
        total_credit_mnt=('credit_mnt', 'sum'),
        closing_balance_mnt=('balance_mnt', 'last')
    ).reset_index()
    acct['account_name'] = acct['account_code'].map(anames)
    acct.insert(0, 'report_year', yr)
    rm = df.groupby(['month', 'account_code', 'counterparty_name']).agg(
        transaction_count=('debit_mnt', 'count'),
        total_debit=('debit_mnt', 'sum'),
        total_credit=('credit_mnt', 'sum'),
    ).reset_index()
    rm['total_amount_mnt'] = rm['total_debit'].abs() + rm['total_credit'].abs()
    rm.insert(0, 'report_year', yr)
    p75a = rm['total_amount_mnt'].quantile(0.75)
    p75c = rm['transaction_count'].quantile(0.75)
    rm['risk_flag_large_txn'] = (rm['total_amount_mnt'] > p75a).astype(int)
    rm['risk_flag_high_frequency'] = (rm['transaction_count'] > p75c).astype(int)
    rm['risk_score'] = rm['risk_flag_large_txn'] + rm['risk_flag_high_frequency']
    rm['account_category'] = rm['account_code'].str[:1].map(
        {'1': 'Хөрөнгө', '2': 'Өр', '3': 'Эздийн өмч', '4': 'Зардал', '5': 'Орлого', '6': 'Орлого', '7': 'Зардал'}
    ).fillna('')
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        monthly.to_excel(w, sheet_name='02_MONTHLY_SUMMARY', index=False)
        acct.to_excel(w, sheet_name='03_ACCOUNT_SUMMARY', index=False)
        rm.to_excel(w, sheet_name='04_RISK_MATRIX', index=False)
    buf.seek(0)
    n_risk = len(rm[rm['risk_score'] > 0])
    return buf, monthly, acct, rm, n_risk

def read_ledger(f):
    raw = f.read()
    f.seek(0)
    if raw[:2] == b'\x1f\x8b':
        return pd.read_csv(io.StringIO(gzip.decompress(raw).decode('utf-8')), dtype={'account_code': str})
    return pd.read_csv(io.BytesIO(raw), dtype={'account_code': str})

def get_year(name):
    for y in range(2020, 2030):
        if str(y) in name:
            return y
    return 2025

def load_tb(files):
    frames = []
    stats = {}
    for f in files:
        year = get_year(f.name)
        df = pd.read_excel(f, sheet_name='02_ACCOUNT_SUMMARY')
        df['year'] = year
        for c in ['turnover_debit', 'turnover_credit', 'closing_debit', 'closing_credit', 'opening_debit', 'opening_credit']:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
        if 'net_change_signed' in df.columns:
            df['net_change_signed'] = pd.to_numeric(df['net_change_signed'], errors='coerce').fillna(0)
        stats[year] = {'accounts': len(df), 'turnover_d': df['turnover_debit'].sum(), 'turnover_c': df['turnover_credit'].sum()}
        frames.append(df)
    return pd.concat(frames, ignore_index=True), stats

def load_ledger_stats(files):
    """Ledger файлуудыг уншиж stats + бүрэн DataFrame буцаана."""
    stats = {}
    all_frames = []
    for f in files:
        year = get_year(f.name)
        f.seek(0)
        df = read_ledger(f)
        for c in ['debit_mnt', 'credit_mnt']:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
        df['report_year'] = str(year)
        mo = df.groupby('month').agg(rows=('debit_mnt', 'count'), debit=('debit_mnt', 'sum'), credit=('credit_mnt', 'sum'))
        stats[year] = {'rows': len(df), 'accounts': df['account_code'].nunique(), 'months': df['month'].nunique(), 'monthly': mo}
        all_frames.append(df)
    full_df = pd.concat(all_frames, ignore_index=True) if all_frames else pd.DataFrame()
    return stats, full_df

def load_part1(files):
    all_rm = []
    all_mo = []
    for f in files:
        year = get_year(f.name)
        try:
            rm = pd.read_excel(f, sheet_name='04_RISK_MATRIX')
            rm['year'] = year
            all_rm.append(rm)
        except Exception:
            pass
        try:
            mo = pd.read_excel(f, sheet_name='02_MONTHLY_SUMMARY')
            mo['year'] = year
            all_mo.append(mo)
        except Exception:
            pass
    rm_all = pd.concat(all_rm, ignore_index=True) if all_rm else pd.DataFrame()
    mo_all = pd.concat(all_mo, ignore_index=True) if all_mo else pd.DataFrame()
    return rm_all, mo_all

def engineer_txn_features(d):
    """Гүйлгээ бүрээс шинж чанар үүсгэнэ. Дутуу багана байвал 0 утга ашиглана."""
    d = d.copy()
    # Баганууд байгаа эсэхийг шалгаж, дутууг нэмэх
    for c in ['debit_mnt','credit_mnt','account_code','account_name','counterparty_name','transaction_description','transaction_date']:
        if c not in d.columns:
            d[c] = '' if c in ('account_code','account_name','counterparty_name','transaction_description','transaction_date') else 0
    d['debit_mnt'] = pd.to_numeric(d['debit_mnt'], errors='coerce').fillna(0)
    d['credit_mnt'] = pd.to_numeric(d['credit_mnt'], errors='coerce').fillna(0)
    d['account_code'] = d['account_code'].astype(str).fillna('000')
    d['account_name'] = d['account_name'].astype(str).fillna('')
    d['counterparty_name'] = d['counterparty_name'].astype(str).fillna('')
    d['transaction_description'] = d['transaction_description'].astype(str).fillna('')
    d['transaction_date'] = d['transaction_date'].astype(str).fillna('')

    d['amount'] = d['debit_mnt'].abs() + d['credit_mnt'].abs()
    d['log_amount'] = np.log1p(d['amount'])
    d['is_debit'] = (d['debit_mnt'] > 0).astype(int)

    # Дансны ангилал
    try:
        le2 = LabelEncoder()
        d['acct_cat_num'] = le2.fit_transform(d['account_code'].str[:3])
    except:
        d['acct_cat_num'] = 0

    # Бенфорд
    digits = d['amount'].apply(lambda x: int(str(int(abs(x)))[0]) if abs(x) >= 1 else 0)
    d['benford_digit'] = digits
    benford_exp = {1:0.301,2:0.176,3:0.125,4:0.097,5:0.079,6:0.067,7:0.058,8:0.051,9:0.046}
    af = d[d['benford_digit']>0]['benford_digit'].value_counts(normalize=True)
    d['benford_dev'] = d['benford_digit'].map(lambda x: abs(af.get(x,0)-benford_exp.get(x,0)) if x>0 else 0)

    # Тэгс тоо
    d['is_round'] = (((d['amount']>=1e6)&(d['amount']%1e6==0)).astype(int) + ((d['amount']>=1e3)&(d['amount']%1e3==0)).astype(int))

    # Данс доторх z-score
    try:
        as2 = d.groupby('account_code')['amount'].agg(['mean','std']).fillna(0)
        as2.columns = ['acct_mean','acct_std']
        d = d.merge(as2, on='account_code', how='left')
        d['amt_zscore'] = np.where(d['acct_std']>0, (d['amount']-d['acct_mean'])/d['acct_std'], 0)
        d['amt_zscore'] = d['amt_zscore'].clip(-10,10).fillna(0)
    except:
        d['acct_mean'] = 0; d['acct_std'] = 0; d['amt_zscore'] = 0

    # Ховор харилцагч
    try:
        cp_f = d['counterparty_name'].value_counts()
        d['cp_rare'] = (d['counterparty_name'].map(cp_f).fillna(0) <= 3).astype(int)
    except:
        d['cp_rare'] = 0

    # Ховор данс-харилцагч хос
    try:
        d['pair'] = d['account_code'] + '|' + d['counterparty_name']
        pf = d['pair'].value_counts()
        d['pair_rare'] = (d['pair'].map(pf).fillna(0) <= 2).astype(int)
    except:
        d['pair_rare'] = 0

    # Тайлбар
    d['desc_empty'] = (d['transaction_description'].str.len() == 0).astype(int)

    # Давхардал
    try:
        d['dup_key'] = d['account_code'] + '|' + d['amount'].astype(str) + '|' + d['transaction_date']
        dk = d['dup_key'].value_counts()
        d['is_dup'] = (d['dup_key'].map(dk).fillna(1) > 1).astype(int)
    except:
        d['is_dup'] = 0

    # Цаг
    d['day'] = pd.to_numeric(d['transaction_date'].str[8:10], errors='coerce').fillna(15)
    d['month_num'] = pd.to_numeric(d['transaction_date'].str[5:7], errors='coerce').fillna(6)
    d['is_month_end'] = (d['day'] >= 28).astype(int)
    d['is_year_end'] = (d['month_num'] == 12).astype(int)

    # ═══ ТАЙЛБАР ↔ ДАНСНЫ НЭР ТУЛГАЛТ ═══
    d['desc_mismatch'] = 0
    d['name_no_overlap'] = 0
    d['dir_mismatch'] = 0
    try:
        stop_w = {'дансны','данс','нийт','бусад','зардал','орлого','төлбөр','хөрөнгө','тооцоо','бүртгэл','дүн','төгрөг','сая','мянга','журнал','гүйлгээ','баримт'}
        # Данс бүрийн ердийн тайлбарын үгс
        acct_words = {}
        for code in d['account_code'].unique():
            all_desc = ' '.join(d.loc[d['account_code']==code, 'transaction_description'].str.lower())
            wc = Counter(re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', all_desc))
            acct_words[code] = set(w for w,c in wc.items() if c >= 3 and len(w) >= 3)

        # Vectorized desc_mismatch
        def _check_mismatch(code, tx_desc):
            tx = str(tx_desc).lower() if tx_desc else ''
            if not tx or code not in acct_words or not acct_words[code]: return 0
            tx_words = set(re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', tx))
            return 0 if len(tx_words & acct_words[code]) > 0 else 1
        d['desc_mismatch'] = [_check_mismatch(c, t) for c, t in zip(d['account_code'], d['transaction_description'])]

        # Vectorized name_no_overlap
        def _extract_kw(text):
            if not text: return set()
            return set(w for w in re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', str(text).lower()) if w not in stop_w and len(w) >= 3)
        def _check_overlap(aname, tdesc):
            nk = _extract_kw(aname)
            dk2 = _extract_kw(tdesc)
            if not nk or not dk2: return 0
            return 0 if len(nk & dk2) > 0 else 1
        d['name_no_overlap'] = [_check_overlap(a, t) for a, t in zip(d['account_name'], d['transaction_description'])]
    except:
        pass

    # Дансны чиглэл зөрчил
    try:
        af2 = d['account_code'].str[0]
        d.loc[(af2=='1')&(d['credit_mnt']>0)&(d['debit_mnt']==0), 'dir_mismatch'] = 1
        d.loc[(af2=='2')&(d['debit_mnt']>0)&(d['credit_mnt']==0), 'dir_mismatch'] = 1
        d.loc[(af2=='5')&(d['debit_mnt']>0)&(d['credit_mnt']==0), 'dir_mismatch'] = 1
        d.loc[(af2.isin(['6','7','8']))&(d['credit_mnt']>0)&(d['debit_mnt']==0), 'dir_mismatch'] = 1
    except:
        pass

    return d

def run_txn_anomaly(df, cont=0.05):
    """Гүйлгээний аномали илрүүлэлт."""
    feats = ['log_amount','acct_cat_num','benford_dev','is_round','amt_zscore','cp_rare','pair_rare',
             'desc_empty','is_month_end','is_year_end','is_dup','is_debit','desc_mismatch','name_no_overlap','dir_mismatch']
    # Бүх feature багана байгаа эсэхийг шалгах
    for f in feats:
        if f not in df.columns:
            df[f] = 0
    X = df[feats].fillna(0).replace([np.inf,-np.inf], 0).astype(float)
    iso = IsolationForest(contamination=cont, random_state=42, n_estimators=200, n_jobs=1)
    df['txn_anomaly'] = (iso.fit_predict(X)==-1).astype(int)
    df['txn_score'] = -iso.score_samples(X)
    try:
        z = np.abs(StandardScaler().fit_transform(X))
        df['txn_zscore_flag'] = (z.max(axis=1)>2.5).astype(int)
    except:
        df['txn_zscore_flag'] = 0
    df['txn_risk'] = (df['txn_anomaly']*3 + df['txn_zscore_flag']*2 + df['is_dup']*2 + df['cp_rare'] +
        df['pair_rare'] + (df['amt_zscore'].abs()>3).astype(int)*2 + df['desc_empty'] +
        df['desc_mismatch']*2 + df['name_no_overlap'] + df['dir_mismatch']*2)
    df['txn_risk_level'] = pd.cut(df['txn_risk'], bins=[-1,3,7,12,100],
        labels=['🟢 Бага','🟡 Дунд','🟠 Өндөр','🔴 Маш өндөр'])
    return df, feats

def run_ml(tb_all, cont, n_est):
    df = tb_all.copy()
    df['cat_code'] = df['account_code'].astype(str).str[:3]
    le = LabelEncoder()
    df['cat_num'] = le.fit_transform(df['cat_code'])
    df['log_turn_d'] = np.log1p(df['turnover_debit'].abs())
    df['log_turn_c'] = np.log1p(df['turnover_credit'].abs())
    df['log_close_d'] = np.log1p(df['closing_debit'].abs())
    df['log_close_c'] = np.log1p(df['closing_credit'].abs())
    df['turn_ratio'] = (df['turnover_debit'] / df['turnover_credit'].replace(0, np.nan)).fillna(0).replace([np.inf, -np.inf], 0)
    if 'net_change_signed' in df.columns:
        df['log_abs_change'] = np.log1p(df['net_change_signed'].abs())
    else:
        df['log_abs_change'] = np.log1p((df['closing_debit'] - df['opening_debit']).abs())
    feats = ['cat_num', 'log_turn_d', 'log_turn_c', 'log_close_d', 'log_close_c', 'turn_ratio', 'log_abs_change', 'year']
    X = df[feats].fillna(0).replace([np.inf, -np.inf], 0)
    iso = IsolationForest(contamination=cont, random_state=42, n_estimators=200)
    df['iso_anomaly'] = (iso.fit_predict(X) == -1).astype(int)
    sc = StandardScaler()
    df['zscore_anomaly'] = (np.abs(sc.fit_transform(X)).max(axis=1) > 2.0).astype(int)
    p95 = df['turn_ratio'].quantile(0.95)
    df['turn_anomaly'] = ((df['turn_ratio'] > p95) | (df['turn_ratio'] < -p95)).astype(int)
    df['ensemble_anomaly'] = ((df['iso_anomaly'] == 1) | ((df['zscore_anomaly'] == 1) & (df['turn_anomaly'] == 1))).astype(int)
    y = df['ensemble_anomaly'].values
    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
    models = {
        'Random Forest': RandomForestClassifier(n_estimators=n_est, max_depth=10, random_state=42, class_weight='balanced'),
        'Gradient Boosting': GradientBoostingClassifier(n_estimators=150, max_depth=5, learning_rate=0.1, random_state=42),
        'Logistic Regression': LogisticRegression(max_iter=1000, random_state=42, class_weight='balanced'),
    }
    res = {}
    for nm, mdl in models.items():
        yp = cross_val_predict(mdl, X, y, cv=cv, method='predict')
        ypr = cross_val_predict(mdl, X, y, cv=cv, method='predict_proba')[:, 1]
        res[nm] = {'pred': yp, 'prob': ypr, 'precision': precision_score(y, yp), 'recall': recall_score(y, yp), 'f1': f1_score(y, yp), 'auc': roc_auc_score(y, ypr)}
    best = max(res, key=lambda k: res[k]['f1'])
    rf = models['Random Forest']
    rf.fit(X, y)
    fi = pd.DataFrame({'feature': feats, 'importance': rf.feature_importances_}).sort_values('importance', ascending=False)
    nt = len(df)
    ns = int(nt * 0.20)
    at = df['turnover_debit'].abs() + df['turnover_credit'].abs()
    wt = at / at.sum()
    wt = wt.fillna(1 / nt)
    np.random.seed(42)
    ms = np.zeros(nt, dtype=int)
    ms[np.random.choice(nt, size=ns, replace=False, p=wt.values)] = 1
    ym = (ms & y).astype(int)
    return df, X, y, feats, res, best, fi, ym

# ═══════════════════════════════════════
# 🧠 УХААЛАГ ФАЙЛ ТАНИХ СИСТЕМ
# ═══════════════════════════════════════
def detect_file_type(f):
    """Файлын төрлийг автоматаар таних. Returns: (type, year)
    Types: 'raw_tb', 'edt', 'tb_std', 'ledger', 'part1', 'unknown'
    """
    name = f.name.lower()
    fname_orig = f.name
    year = get_year(f.name)

    # CSV/GZ → Ledger
    if name.endswith('.csv') or name.endswith('.gz') or name.endswith('.csv.gz'):
        return 'ledger', year

    # XLSX → need to check
    if not name.endswith('.xlsx'):
        return 'unknown', year

    # ── Файлын нэрээр хурдан таних ──
    name_check = fname_orig.lower().replace('_', ' ').replace('-', ' ')
    # ЕДТ / Ерөнхий журнал / Journal
    edt_keywords = ['ерөнхий журнал', 'ерөнхий дэвтэр', 'едт', 'edt', 'general ledger', 'general journal',
                    'еренхий журнал', 'journal gc', 'journal entry', 'journal entries']
    for kw in edt_keywords:
        if kw in name_check:
            return 'edt', year
    # ГҮЙЛГЭЭ_БАЛАНС / Trial Balance / Journal TB
    tb_keywords = ['гүйлгээ баланс', 'гүйлгээ_баланс', 'гуйлгээ баланс', 'trial balance',
                   'гүйлгэ баланс', 'гуйлгэ баланс', 'journal, tb', 'journal tb']
    for kw in tb_keywords:
        if kw in name_check:
            return 'raw_tb', year
    # TB_standardized
    if 'tb_standardized' in name_check or 'tb standardized' in name_check:
        return 'tb_std', year
    # Part1
    if 'part1' in name_check or 'part 1' in name_check:
        return 'part1', year
    # Ledger
    if 'ledger' in name_check or 'prototype_ledger' in name_check:
        return 'ledger', year

    # ── Sheet бүтцээр таних ──
    import openpyxl
    try:
        raw = f.read()
        f.seek(0)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        sheets = wb.sheetnames

        # TB_standardized: has '02_ACCOUNT_SUMMARY' sheet
        if '02_ACCOUNT_SUMMARY' in sheets:
            if '04_RISK_MATRIX' in sheets:
                wb.close()
                return 'part1', year
            wb.close()
            return 'tb_std', year

        # Part1: has '04_RISK_MATRIX' sheet
        if '04_RISK_MATRIX' in sheets:
            wb.close()
            return 'part1', year

        # ── Агуулгаар таних (200 мөр хүртэл шалгана) ──
        ws = wb[sheets[0]]
        sample_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            sample_rows.append(row)
            if i >= 200:
                break
        wb.close()

        # ЕДТ: contains "Данс:" or "Компани:" or "ЕРӨНХИЙ" or "Журнал:" pattern
        for row in sample_rows:
            if row[0] is not None:
                s = str(row[0]).strip()
                if s.startswith('Данс:') or s.startswith('Компани:') or s.startswith('ЕРӨНХИЙ') or s.startswith('Журнал:'):
                    return 'edt', year
            # ЕДТ: column with "Данс:" might be in other columns too
            for cell in row[:5]:
                if cell is not None and 'Данс:' in str(cell):
                    return 'edt', year

        # Хүснэгтэн журнал: Огноо + Дебет + Кредит + Мөнгөн дүн
        header_vals = [str(c).strip().lower() for c in sample_rows[0] if c is not None] if sample_rows else []
        has_journal_header = (
            any('огноо' in v or 'date' in v for v in header_vals) and
            any('дебет' in v or 'debit' in v for v in header_vals) and
            any('кредит' in v or 'credit' in v for v in header_vals) and
            any('мөнгөн дүн' in v or 'amount' in v for v in header_vals)
        )
        if has_journal_header:
            return 'edt', year

        # ГҮЙЛГЭЭ_БАЛАНС: has account codes like 101-XX-XX-XXX in column B
        for row in sample_rows:
            if len(row) >= 2 and row[1] is not None:
                code = str(row[1]).strip()
                if re.match(r'\d{3}-\d{2}-\d{2}-\d{3}', code):
                    return 'raw_tb', year

        # Fallback: check if it looks like a balance sheet
        for row in sample_rows:
            if row[0] is not None:
                try:
                    int(float(row[0]))
                    if len(row) >= 8 and row[1] is not None and re.match(r'\d{3}-', str(row[1])):
                        return 'raw_tb', year
                except:
                    pass

        return 'unknown', year
    except Exception:
        f.seek(0)
        return 'unknown', year

FILE_TYPE_LABELS = {
    'raw_tb': ('📗 ГҮЙЛГЭЭ_БАЛАНС', 'Гүйлгээ-балансын түүхий файл → TB болгон хөрвүүлнэ'),
    'edt': ('📘 ЕДТ', 'Ерөнхий дэвтрийн тайлан → Ledger + Part1 болгон хөрвүүлнэ'),
    'tb_std': ('📊 TB_standardized', 'Стандартчилсан гүйлгээ-баланс → Шинжилгээнд бэлэн'),
    'ledger': ('📄 Ledger CSV/GZ', 'Ерөнхий дэвтрийн гүйлгээ → Шинжилгээнд бэлэн'),
    'part1': ('📈 Part1', 'Сарын нэгтгэл + Эрсдэлийн матриц → Шинжилгээнд бэлэн'),
    'unknown': ('❓ Тодорхойгүй', 'Файлын төрлийг таних боломжгүй'),
}
if page.startswith("1"):
    st.header("1️⃣ Өгөгдөл бэлтгэх")
    st.markdown("""
    <div style="background-color: #E3F2FD; padding: 15px; border-radius: 8px; border-left: 4px solid #1565C0; margin-bottom: 15px;">
        <b>📂 Ямар ч файлыг оруулаарай!</b> Систем автоматаар таниж, зөв формат руу хөрвүүлнэ.<br>
        <span style="color: #555; font-size: 13px;">
        Дэмжих файлууд: ГҮЙЛГЭЭ_БАЛАНС (.xlsx), ЕДТ (.xlsx) — хэдэн ч файл, ямар ч дараалал
        </span>
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader("📎 Бүх файлуудаа энд оруулна уу", type=['xlsx', 'csv', 'gz'], accept_multiple_files=True, key='smart_prep')

    if uploaded:
        detected = []
        for f in uploaded:
            ftype, year = detect_file_type(f)
            f.seek(0)
            detected.append({'file': f, 'type': ftype, 'year': year, 'name': f.name})

        st.markdown("### 🔍 Таних үр дүн")
        det_rows = []
        for d in detected:
            label, desc = FILE_TYPE_LABELS.get(d['type'], FILE_TYPE_LABELS['unknown'])
            det_rows.append({'Файл': d['name'], 'Төрөл': label, 'Он': d['year'], 'Тайлбар': desc})
        st.dataframe(pd.DataFrame(det_rows), use_container_width=True, hide_index=True)

        raw_tbs = [d for d in detected if d['type'] == 'raw_tb']
        edts = [d for d in detected if d['type'] == 'edt']
        unknowns = [d for d in detected if d['type'] == 'unknown']
        ready = [d for d in detected if d['type'] in ('tb_std', 'ledger', 'part1')]

        if unknowns:
            st.warning(f"⚠️ {len(unknowns)} файл танигдсангүй: {', '.join(u['name'] for u in unknowns)}")
        if ready:
            st.success(f"✅ Шинжилгээнд бэлэн: {len(ready)} файл → **2️⃣ Шинжилгээ** руу шилжээрэй")

        if raw_tbs or edts:
            if st.button("⚙️ Хөрвүүлэлт эхлүүлэх", type="primary", use_container_width=True, key='btn_smart'):
                if raw_tbs:
                    if 'tb_res' not in st.session_state:
                        st.session_state.tb_res = {}
                    for d in raw_tbs:
                        with st.spinner(f"📗 ГҮЙЛГЭЭ_БАЛАНС {d['year']} хөрвүүлж байна..."):
                            d['file'].seek(0)
                            buf, tb_s = process_raw_tb(d['file'])
                            st.session_state.tb_res[d['year']] = {'buf': buf.getvalue(), 'tb': tb_s}
                        st.success(f"✅ TB {d['year']}: {len(tb_s):,} данс")
                if edts:
                    if 'led_res' not in st.session_state:
                        st.session_state.led_res = {}
                    edt_by_year = {}
                    for d in edts:
                        edt_by_year.setdefault(d['year'], []).append(d['file'])
                    for yr in sorted(edt_by_year):
                        with st.spinner(f"📘 ЕДТ {yr} хөрвүүлж байна ({len(edt_by_year[yr])} файл)..."):
                            frames = []
                            for f in edt_by_year[yr]:
                                f.seek(0)
                                df_e, cnt_e = process_edt(f, yr)
                                if cnt_e > 0:
                                    frames.append(df_e)
                            if frames:
                                st.session_state.led_res[yr] = pd.concat(frames, ignore_index=True)
                                st.success(f"✅ ЕДТ {yr}: {len(st.session_state.led_res[yr]):,} гүйлгээ")
                            else:
                                st.warning(f"⚠️ {yr} оны ЕДТ файл(уудаас) гүйлгээ уншигдсангүй. Файлын формат шалгана уу.")

    if 'tb_res' in st.session_state and st.session_state.tb_res:
        st.markdown("---\n### 📥 TB файлууд")
        for yr in sorted(st.session_state.tb_res):
            d = st.session_state.tb_res[yr]
            st.download_button(f"📥 TB_standardized_{yr}.xlsx ({len(d['tb']):,} данс)", d['buf'], f"TB_standardized_{yr}1231.xlsx", key=f"dtb{yr}")

    if 'led_res' in st.session_state and st.session_state.led_res:
        st.markdown("---\n### 📥 Ledger + Part1")
        cols_out = ['report_year','account_code','account_name','transaction_no','transaction_date','journal_no','document_no','counterparty_name','counterparty_id','transaction_description','debit_mnt','credit_mnt','balance_mnt','month']
        for yr in sorted(st.session_state.led_res):
            dfy = st.session_state.led_res[yr]
            if dfy.empty or 'debit_mnt' not in dfy.columns:
                st.warning(f"⚠️ {yr} оны ЕДТ файлаас гүйлгээ уншигдсангүй. Файлын формат тохирохгүй байж магадгүй.")
                continue
            dfy['debit_mnt'] = pd.to_numeric(dfy['debit_mnt'], errors='coerce').fillna(0)
            dfy['credit_mnt'] = pd.to_numeric(dfy['credit_mnt'], errors='coerce').fillna(0)
            with st.expander(f"📅 {yr} — {len(dfy):,} гүйлгээ", expanded=True):
                p1_buf, p1_mo, p1_acct, p1_rm, n_risk = generate_part1(dfy, yr)
                c1x, c2x, c3x = st.columns(3)
                c1x.metric("Гүйлгээ", f"{len(dfy):,}")
                c2x.metric("Эрсдэлийн хос", f"{len(p1_rm):,}")
                c3x.metric("Эрсдэлтэй", f"{n_risk:,}")
                gz_bytes = gzip.compress(dfy[cols_out].to_csv(index=False).encode('utf-8'))
                st.download_button(f"📥 ledger_{yr}.csv.gz", gz_bytes, f"prototype_ledger_{yr}.csv.gz", key=f"dled{yr}")
                st.download_button(f"📥 part1_{yr}.xlsx", p1_buf.getvalue(), f"prototype_part1_{yr}.xlsx", key=f"dp1{yr}")

# ═══════════════════════════════════════
# 2️⃣ ШИНЖИЛГЭЭ
# ═══════════════════════════════════════
elif page.startswith("2"):
    st.header("2️⃣ ХОУ Шинжилгээ")
    st.markdown("""
    <div style="background-color: #E8F5E9; padding: 15px; border-radius: 8px; border-left: 4px solid #2E7D32; margin-bottom: 15px;">
        <b>📂 Ямар ч файлаа нэг дор оруулаарай!</b> Систем автоматаар таниж, хөрвүүлж, шинжилгээг ажиллуулна.<br>
        <span style="color: #555; font-size: 13px;">
        ГҮЙЛГЭЭ_БАЛАНС, ЕДТ, TB, Ledger, Part1 — бүгдийг нь оруулаад болно. Систем өөрөө ялгана.
        </span>
    </div>
    """, unsafe_allow_html=True)

    all_files = st.file_uploader("📎 Бүх файлуудаа энд оруулна уу (ямар ч формат, хэдэн ч файл)", type=['xlsx', 'csv', 'gz'], accept_multiple_files=True, key='smart_analysis')

    tb_files = []
    led_files = []
    p1_files = []

    if all_files:
        detected = []
        for f in all_files:
            ftype, year = detect_file_type(f)
            f.seek(0)
            detected.append({'file': f, 'type': ftype, 'year': year, 'name': f.name})

        # Show detection summary
        det_rows = []
        for d in detected:
            label, desc = FILE_TYPE_LABELS.get(d['type'], FILE_TYPE_LABELS['unknown'])
            det_rows.append({'Файл': d['name'], 'Төрөл': label, 'Он': d['year'], 'Тайлбар': desc})
        st.dataframe(pd.DataFrame(det_rows), use_container_width=True, hide_index=True)

        # Auto-convert raw files + route ready files
        raw_tbs = [d for d in detected if d['type'] == 'raw_tb']
        edts = [d for d in detected if d['type'] == 'edt']
        need_convert = len(raw_tbs) > 0 or len(edts) > 0

        if need_convert:
            st.info(f"🔄 **{len(raw_tbs)} ГҮЙЛГЭЭ_БАЛАНС + {len(edts)} ЕДТ** файл автоматаар хөрвүүлэгдэнэ.")

        for d in detected:
            if d['type'] == 'tb_std':
                tb_files.append(d['file'])
            elif d['type'] == 'ledger':
                led_files.append(d['file'])
            elif d['type'] == 'part1':
                p1_files.append(d['file'])
            elif d['type'] == 'raw_tb':
                # Auto-convert ГҮЙЛГЭЭ_БАЛАНС → TB_standardized
                with st.spinner(f"📗 {d['name']} → TB хөрвүүлж байна..."):
                    d['file'].seek(0)
                    buf, _ = process_raw_tb(d['file'])
                    buf.seek(0)
                    tb_wrap = io.BytesIO(buf.getvalue())
                    tb_wrap.name = f"TB_standardized_{d['year']}1231.xlsx"
                    tb_files.append(tb_wrap)
                st.success(f"✅ {d['name']} → TB хөрвүүлсэн")
            elif d['type'] == 'edt':
                # Auto-convert ЕДТ → Ledger + Part1
                with st.spinner(f"📘 {d['name']} → Ledger + Part1 хөрвүүлж байна..."):
                    d['file'].seek(0)
                    df_edt, cnt = process_edt(d['file'], d['year'])
                if cnt == 0 or df_edt.empty:
                    st.warning(f"⚠️ **{d['name']}** — ЕДТ гэж танигдсан ч гүйлгээ уншигдсангүй. Файлын формат тохирохгүй байж магадгүй.")
                else:
                    cols_out = ['report_year','account_code','account_name','transaction_no','transaction_date',
                                'journal_no','document_no','counterparty_name','counterparty_id',
                                'transaction_description','debit_mnt','credit_mnt','balance_mnt','month']
                    df_edt['debit_mnt'] = pd.to_numeric(df_edt['debit_mnt'], errors='coerce').fillna(0)
                    df_edt['credit_mnt'] = pd.to_numeric(df_edt['credit_mnt'], errors='coerce').fillna(0)
                    csv_bytes = df_edt[cols_out].to_csv(index=False).encode('utf-8')
                    led_wrap = io.BytesIO(csv_bytes)
                    led_wrap.name = f"prototype_ledger_{d['year']}.csv"
                    led_files.append(led_wrap)
                    # ЕДТ DataFrame-ийг шууд хадгалах (гүйлгээний шинжилгээнд ашиглана)
                    if 'edt_frames' not in st.session_state:
                        st.session_state['edt_frames'] = []
                    st.session_state['edt_frames'].append(df_edt)
                    p1_buf, _, _, _, _ = generate_part1(df_edt, d['year'])
                    p1_buf.seek(0)
                    p1_wrap = io.BytesIO(p1_buf.getvalue())
                    p1_wrap.name = f"prototype_part1_{d['year']}.xlsx"
                    p1_files.append(p1_wrap)
                    st.success(f"✅ {d['name']} → Ledger ({cnt:,} гүйлгээ) + Part1 хөрвүүлсэн")

        if tb_files and led_files:
            st.success(f"🎯 Бүрэн шинжилгээнд бэлэн: TB {len(tb_files)} | Ledger {len(led_files)} | Part1 {len(p1_files)}")
        elif led_files and not tb_files:
            st.success(f"🎯 Гүйлгээний шинжилгээнд бэлэн: Ledger {len(led_files)} файл (TB нэмбэл дансны шинжилгээ ч ажиллана)")
        elif tb_files and not led_files:
            st.info("👆 Ledger (.csv/.gz) эсвэл ЕДТ (.xlsx) файл нэмнэ үү")
        elif not tb_files and not led_files and not need_convert:
            st.info("👆 TB + Ledger файлуудаа оруулна уу")

    st.markdown("""
    <div style="background:#F5F5F5; padding:12px; border-radius:8px; margin-bottom:10px;">
    <b>⚙️ Шинжилгээний тохиргоо</b>
    </div>
    """, unsafe_allow_html=True)
    c1s, c2s = st.columns(2)
    with c1s:
        cont = st.slider("🎯 IF contamination — Тусгаарлалтын ойн хэвийн бус хувь", 0.05, 0.20, 0.10, 0.01,
            help="Isolation Forest (Тусгаарлалтын ой) алгоритм нийт дансны хэдэн хувийг хэвийн бус гэж үзэх. "
                 "0.05 (5%) = зөвхөн хамгийн сэжигтэй 5%. "
                 "0.10 (10%) = 10% илрүүлнэ (анхдагч). "
                 "0.20 (20%) = илүү өргөн хүрээтэй шалгана.")
    with c2s:
        nest = st.slider("🌲 RF n_estimators — Санамсаргүй ойн модны тоо", 50, 500, 200, 50,
            help="Random Forest (Санамсаргүй ой) загварын модны тоо. "
                 "50 = хурдан, бага нарийвчлал. "
                 "200 = тэнцвэртэй (анхдагч). "
                 "500 = удаан, өндөр нарийвчлал.")

    has_any = tb_files or led_files
    if st.button("🚀 Шинжилгээ", type="primary", use_container_width=True) and has_any:
        # Дансны түвшний шинжилгээ (TB + Ledger хоёулаа байвал)
        df = pd.DataFrame(); X = np.array([]); y = np.array([]); feats = []
        res = {}; best = ''; fi = pd.DataFrame(); ym = np.array([])
        tb_st = {}; led_st = {}; ledger_full = pd.DataFrame()
        rm_all = pd.DataFrame(); mo_all = pd.DataFrame()

        if tb_files and led_files:
            with st.spinner("TB уншиж байна..."):
                tb_all, tb_st = load_tb(tb_files)
            with st.spinner("Ledger уншиж байна..."):
                led_st, ledger_full = load_ledger_stats(led_files)
            if p1_files:
                with st.spinner("Part1 уншиж байна..."):
                    rm_all, mo_all = load_part1(p1_files)
            with st.spinner("🤖 Дансны түвшний шинжилгээ..."):
                df, X, y, feats, res, best, fi, ym = run_ml(tb_all, cont, nest)
        elif led_files:
            # Зөвхөн Ledger байвал — stats + full уншна
            with st.spinner("Ledger уншиж байна..."):
                led_st, ledger_full = load_ledger_stats(led_files)
            if p1_files:
                with st.spinner("Part1 уншиж байна..."):
                    rm_all, mo_all = load_part1(p1_files)

        # Гүйлгээний түвшний шинжилгээ (Ledger эсвэл ЕДТ байвал)
        txn_result = pd.DataFrame()
        # ЕДТ-ээс шууд хадгалсан DataFrame-үүдийг нэгтгэх
        edt_frames = st.session_state.get('edt_frames', [])
        all_txn_frames = []
        if len(ledger_full) > 0:
            all_txn_frames.append(ledger_full)
        elif edt_frames:
            all_txn_frames.extend(edt_frames)

        if all_txn_frames:
            txn_combined = pd.concat(all_txn_frames, ignore_index=True)
            with st.spinner(f"🔍 Гүйлгээний түвшний шинжилгээ ({len(txn_combined):,} гүйлгээ)..."):
                try:
                    sample_n = min(len(txn_combined), 50000)
                    txn_s = txn_combined.sample(n=sample_n, random_state=42) if len(txn_combined) > sample_n else txn_combined.copy()
                    txn_s = engineer_txn_features(txn_s)
                    txn_result, _ = run_txn_anomaly(txn_s, cont)
                except Exception as e:
                    st.warning(f"⚠️ Гүйлгээний шинжилгээ алдаа: {e}")
        # Store all results in session_state
        st.session_state['analysis_done'] = True
        st.session_state['df'] = df
        st.session_state['X'] = X
        st.session_state['y'] = y
        st.session_state['feats'] = feats
        st.session_state['res'] = res
        st.session_state['best'] = best
        st.session_state['fi'] = fi
        st.session_state['ym'] = ym
        st.session_state['tb_st'] = tb_st
        st.session_state['led_st'] = led_st
        st.session_state['rm_all'] = rm_all
        st.session_state['mo_all'] = mo_all
        st.session_state['txn_result'] = txn_result

    # Display results from session_state (persists across reruns)
    if st.session_state.get('analysis_done', False):
        df = st.session_state['df']
        X = st.session_state['X']
        y = st.session_state['y']
        feats = st.session_state['feats']
        res = st.session_state['res']
        best = st.session_state['best']
        fi = st.session_state['fi']
        ym = st.session_state['ym']
        tb_st = st.session_state['tb_st']
        led_st = st.session_state['led_st']
        rm_all = st.session_state['rm_all']
        mo_all = st.session_state['mo_all']
        txn_result = st.session_state.get('txn_result', pd.DataFrame())

        has_account = len(df) > 0 and len(res) > 0
        has_rm = len(rm_all) > 0
        has_mo = len(mo_all) > 0
        has_txn = len(txn_result) > 0
        n_led = sum(d['rows'] for d in led_st.values()) if led_st else (len(txn_result) if has_txn else 0)

        if has_account:
            st.success(f"✅ {len(df):,} данс, {n_led:,} гүйлгээ шинжлэгдсэн")
        elif has_txn:
            st.success(f"✅ {n_led:,} гүйлгээ шинжлэгдсэн (гүйлгээний түвшний шинжилгээ)")
        
        yrs = sorted(tb_st.keys()) if tb_st else []
        bp = res[best]['pred'] if has_account and best else np.array([])

        tab_names = ["📊 Нэгтгэл", "🔍 Хэвийн бус данс", "⚖️ ХОУ ↔ Уламжлалт", "🧠 Тайлбарлагдах ХОУ", "📋 Жагсаалт"]
        if has_txn:
            tab_names.append("🔴 Гүйлгээний эрсдэл")
            tab_names.append("👤 Харилцагчаар")
        if has_rm:
            tab_names.append("🎯 Эрсдэлийн матриц")
        if has_mo:
            tab_names.append("📈 Сарын хандлага")

        all_tabs = st.tabs(tab_names)

        with all_tabs[0]:
            if not has_account:
                st.info("📊 TB + Ledger файлуудыг оруулахад дансны түвшний шинжилгээ идэвхжинэ. ЕДТ файлаар зөвхөн гүйлгээний шинжилгээ ажиллана.")
            else:
                td.show_summary_description(n_accounts=len(df), n_transactions=n_led, n_risk_pairs=len(rm_all) if has_rm else 0)
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Данс", f"{len(df):,}")
                m2.metric("Гүйлгээ", f"{sum(d['rows'] for d in led_st.values()):,}")
                m3.metric("Аномали", f"{df['ensemble_anomaly'].sum():,} ({df['ensemble_anomaly'].mean()*100:.1f}%)")
                m4.metric("Шилдэг", f"{best} F1={res[best]['f1']:.4f}")
                if has_rm:
                    mr1, mr2 = st.columns(2)
                    mr1.metric("Эрсдэлийн хос", f"{len(rm_all):,}")
                    mr2.metric("Эрсдэлтэй (score>0)", f"{len(rm_all[rm_all['risk_score']>0]):,}")
                fg = make_subplots(rows=1, cols=3, subplot_titles=("Данс", "Эргэлт (T₮)", "ЕДТ мөр"))
                cl3 = ['#2196F3', '#4CAF50', '#FF9800']
                for i, yv in enumerate(yrs):
                    fg.add_trace(go.Bar(x=[str(yv)], y=[tb_st[yv]['accounts']], marker_color=cl3[i % 3], showlegend=False), row=1, col=1)
                    fg.add_trace(go.Bar(x=[str(yv)], y=[tb_st[yv]['turnover_d'] / 1e9], marker_color=cl3[i % 3], showlegend=False), row=1, col=2)
                    if yv in led_st:
                        fg.add_trace(go.Bar(x=[str(yv)], y=[led_st[yv]['rows']], marker_color=cl3[i % 3], showlegend=False), row=1, col=3)
                fg.update_layout(height=350)
                st.plotly_chart(fg, use_container_width=True)
                td.show_summary_interpretation()

        with all_tabs[1]:
          if not has_account:
            st.info("TB + Ledger файл шаардлагатай")
          else:
            td.show_anomaly_description()
            mt = {'Isolation Forest': 'iso_anomaly', 'Z-score': 'zscore_anomaly', 'Turn ratio': 'turn_anomaly', 'ENSEMBLE': 'ensemble_anomaly'}
            ad = []
            for m, c in mt.items():
                row_d = {'Арга': m, 'Нийт': int(df[c].sum())}
                for yv in yrs:
                    mask = df['year'] == yv
                    cnt = df.loc[mask, c].sum()
                    pct = cnt / mask.sum() * 100
                    row_d[str(yv)] = f"{int(cnt)} ({pct:.1f}%)"
                ad.append(row_d)
            st.dataframe(pd.DataFrame(ad), use_container_width=True, hide_index=True)
            st.plotly_chart(px.scatter(df, x='log_turn_d', y='log_abs_change', color=df['ensemble_anomaly'].map({0: 'Хэвийн', 1: 'Аномали'}), facet_col='year', opacity=0.5, color_discrete_map={'Хэвийн': '#90caf9', 'Аномали': '#c62828'}, height=400), use_container_width=True)
            td.show_anomaly_interpretation(
                n_if=int(df['iso_anomaly'].sum()),
                n_zscore=int(df['zscore_anomaly'].sum()),
                n_turn=int(df['turn_anomaly'].sum()),
                n_ensemble=int(df['ensemble_anomaly'].sum())
            )

        with all_tabs[2]:
          if not has_account:
            st.info("TB + Ledger файл шаардлагатай")
          else:
            td.show_ai_vs_mus_description()
            st.dataframe(pd.DataFrame([{'Загвар': n, 'Precision': f"{r['precision']:.4f}", 'Recall': f"{r['recall']:.4f}", 'F1': f"{r['f1']:.4f}", 'AUC': f"{r['auc']:.4f}"} for n, r in res.items()]), use_container_width=True, hide_index=True)
            fg2 = go.Figure()
            for n, r in res.items():
                fpr, tpr, _ = roc_curve(y, r['prob'])
                fg2.add_trace(go.Scatter(x=fpr, y=tpr, name=f"{n} (AUC={r['auc']:.4f})"))
            fg2.add_trace(go.Scatter(x=[0, 1], y=[0, 1], name='Random', line=dict(dash='dash', color='gray')))
            fg2.update_layout(title='ROC Curve', height=400)
            st.plotly_chart(fg2, use_container_width=True)
            st.subheader("Detection Risk")
            dr = []
            for yv in yrs:
                mk = (df['year'] == yv).values
                yt = y[mk]
                nt2 = yt.sum()
                if nt2 > 0:
                    a2 = 1 - (bp[mk] & yt).sum() / nt2
                    m2x = 1 - (ym[mk] & yt).sum() / nt2
                else:
                    a2 = 0
                    m2x = 0
                dr.append({'Жил': yv, 'ХОУ': f"{a2:.4f}", 'MUS 20%': f"{m2x:.4f}", 'Сайжрал': f"{m2x - a2:.4f}"})
            st.dataframe(pd.DataFrame(dr), use_container_width=True, hide_index=True)
            td.show_ai_vs_mus_interpretation(
                rf_f1=f"{res[best]['f1']:.4f}",
                rf_auc=f"{res[best]['auc']:.4f}",
                dr_ai=dr[0]['ХОУ'] if dr else "",
                dr_mus=dr[0]['MUS 20%'] if dr else "",
                mcnemar_chi2="p<0.001"
            )

        with all_tabs[3]:
          if not has_account:
            st.info("TB + Ledger файл шаардлагатай")
          else:
            td.show_xai_description()
            st.plotly_chart(px.bar(fi, x='importance', y='feature', orientation='h', color='importance', color_continuous_scale='Blues', title='Feature Importance').update_layout(height=400, yaxis={'categoryorder': 'total ascending'}), use_container_width=True)
            fi_dict = dict(zip(fi['feature'], fi['importance']))
            td.show_xai_feature_details(feature_importances=fi_dict)
            td.show_xai_interpretation()

        with all_tabs[4]:
          if not has_account:
            st.info("TB + Ledger файл шаардлагатай")
          else:
            td.show_list_description()
            adf = df[df['ensemble_anomaly'] == 1][['year', 'account_code', 'account_name', 'turnover_debit', 'turnover_credit', 'turn_ratio', 'log_abs_change']].copy()
            yf = st.selectbox("Жил", ['Бүгд'] + [str(y2) for y2 in yrs])
            if yf != 'Бүгд':
                adf = adf[adf['year'] == int(yf)]
            st.write(f"Нийт: {len(adf)}")
            st.dataframe(adf, use_container_width=True, hide_index=True, height=500)
            st.download_button("📥 CSV", adf.to_csv(index=False).encode('utf-8-sig'), "anomaly.csv", "text/csv")
            td.show_list_interpretation(n_anomalies=len(adf))

        # ── Гүйлгээний түвшний шинжилгээний табууд ──
        next_idx = 5
        if has_txn:
            with all_tabs[next_idx]:
                st.subheader("🔴 Гүйлгээний түвшний хэвийн бус байдал")
                st.markdown("""
                <div style="background:#fce4ec; padding:12px; border-radius:8px; border-left:4px solid #c62828; margin-bottom:15px;">
                <b>16 шинж чанараар</b> гүйлгээ бүрийг шинжилж хэвийн бус гүйлгээг илрүүлсэн.
                Дансны нэр, гүйлгээний тайлбар, харилцагч, дүн, цаг хугацаа бүгдийг тулган шалгана.
                </div>
                """, unsafe_allow_html=True)

                with st.expander("📋 16 шинж чанарын тайлбар — Юу юуг илрүүлдэг вэ?", expanded=False):
                    st.markdown("""
| # | Шинж чанар | Юу илрүүлдэг | ISA | Эрсдэлийн оноо |
|---|-----------|-------------|-----|----------------|
| 1 | **Данс доторх хэвийн бус дүн** (`amt_zscore`) | Тухайн дансны дундажаас хэт зөрсөн гүйлгээ | ISA 520 | >3σ → +2 |
| 2 | **Бенфордын хуулийн хазайлт** (`benford_dev`) | Эхний оронгийн тархалт зөрсөн → тоон манипуляцийн шинж | ISA 240 | IF-д орно |
| 3 | **Тэгс тоо** (`is_round`) | Бүхэл/тэгс дүнтэй сэжигтэй гүйлгээ (1сая, 10сая) | ISA 240 | 1сая+ → +1 |
| 4 | **Ховор харилцагч** (`cp_rare`) | ≤3 удаа гарсан шинэ/сэжигтэй харилцагч | ISA 550 | +1 |
| 5 | **Ховор данс×харилцагч хос** (`pair_rare`) | Тухайн данс + тухайн харилцагч хослол ер бусын | ISA 550 | +1 |
| 6 | **Давхардсан гүйлгээ** (`is_dup`) | Ижил данс + ижил дүн + ижил огноо | ISA 240 | +2 |
| 7 | **Тайлбаргүй гүйлгээ** (`desc_empty`) | Хоосон тайлбартай гүйлгээ | ISA 500 | +1 |
| 8 | **Сарын эцэс** (`is_month_end`) | Сарын 28+ өдөрт хийсэн гүйлгээ | ISA 240 | IF-д орно |
| 9 | **Жилийн эцэс** (`is_year_end`) | 12-р сарын гүйлгээ | ISA 240 | IF-д орно |
| 10 | **Дансны ангилал** (`acct_cat_num`) | Тодорхой ангиллын дансны эрсдэл өөр | ISA 315 | IF-д орно |
| 11 | **Дебит/кредит чиглэл** (`is_debit`) | Гүйлгээний чиглэл | ISA 240 | IF-д орно |
| 12 | **Гүйлгээний дүн** (`log_amount`) | Том дүнтэй гүйлгээний эрсдэл | ISA 320 | IF-д орно |
| — | | **Тайлбар ↔ Дансны нэр тулгалт (шинэ):** | | |
| 13 | **⚠️ Тайлбар ↔ дансны хэв маяг** (`desc_mismatch`) | Тухайн дансны ердийн тайлбараас зөрсөн | ISA 500 | +2 |
| 14 | **⚠️ Дансны нэр ↔ тайлбар** (`name_no_overlap`) | Дансны нэрийн түлхүүр үг тайлбарт огт байхгүй | ISA 500 | +1 |
| 15 | **⚠️ Дансны төрөл ↔ чиглэл** (`dir_mismatch`) | Хөрөнгийн данс→кредит, орлогын данс→дебит гэх мэт | ISA 240 | +2 |
| 16 | **Гүйлгээний дүн** (`log_amount`) | Нийт гүйлгээний хэмжээ (логарифм масштаб) | ISA 320 | IF-д орно |

**Эрсдэлийн оноо:** 🟢 Бага (0-3) → 🟡 Дунд (4-7) → 🟠 Өндөр (8-12) → 🔴 Маш өндөр (13+)

**IF-д орно** = Isolation Forest алгоритмд шууд шинж чанар болж ордог (тоон оноогүй ч нөлөөлнө)
                    """)

                n_txn_anom = txn_result['txn_anomaly'].sum()
                c1,c2,c3,c4 = st.columns(4)
                c1.metric("Шинжилсэн гүйлгээ", f"{len(txn_result):,}")
                c2.metric("Хэвийн бус", f"{n_txn_anom:,}", delta=f"{n_txn_anom/len(txn_result)*100:.1f}%", delta_color="inverse")
                c3.metric("Тайлбар зөрчилтэй", f"{txn_result['desc_mismatch'].sum():,}")
                c4.metric("Чиглэл зөрсөн", f"{txn_result['dir_mismatch'].sum():,}")

                # Нэмэлт metric row
                c5,c6,c7,c8 = st.columns(4)
                c5.metric("Давхардсан", f"{txn_result['is_dup'].sum():,}")
                c6.metric("Ховор харилцагч", f"{txn_result['cp_rare'].sum():,}")
                c7.metric("Тайлбаргүй", f"{txn_result['desc_empty'].sum():,}")
                c8.metric("Дүн хэт зөрсөн (>3σ)", f"{(txn_result['amt_zscore'].abs()>3).sum():,}")

                # Эрсдэлийн түвшний тархалт
                rl = txn_result['txn_risk_level'].value_counts().reindex(['🟢 Бага','🟡 Дунд','🟠 Өндөр','🔴 Маш өндөр']).fillna(0)
                st.plotly_chart(px.bar(x=rl.index, y=rl.values, color=rl.index, color_discrete_map={'🟢 Бага':'#4CAF50','🟡 Дунд':'#FFC107','🟠 Өндөр':'#FF9800','🔴 Маш өндөр':'#F44336'}, labels={'x':'Эрсдэлийн түвшин','y':'Тоо'}, title="Гүйлгээний эрсдэлийн тархалт").update_layout(height=300, showlegend=False), use_container_width=True)

                # Шинж чанар бүрийн илрүүлэлтийн тойм
                with st.expander("📊 Шинж чанар бүрийн илрүүлэлтийн тойм", expanded=False):
                    feat_summary = []
                    feat_info = [
                        ('amt_zscore', 'Данс доторх хэвийн бус дүн (>3σ)', lambda d: (d['amt_zscore'].abs()>3).sum()),
                        ('benford_dev', 'Бенфордын хазайлт (>0.05)', lambda d: (d['benford_dev']>0.05).sum()),
                        ('is_round', 'Тэгс тоо (1000+)', lambda d: (d['is_round']>0).sum()),
                        ('cp_rare', 'Ховор харилцагч (≤3 удаа)', lambda d: d['cp_rare'].sum()),
                        ('pair_rare', 'Ховор данс×харилцагч хос', lambda d: d['pair_rare'].sum()),
                        ('is_dup', 'Давхардсан гүйлгээ', lambda d: d['is_dup'].sum()),
                        ('desc_empty', 'Тайлбаргүй гүйлгээ', lambda d: d['desc_empty'].sum()),
                        ('is_month_end', 'Сарын эцэс (28+)', lambda d: d['is_month_end'].sum()),
                        ('is_year_end', 'Жилийн эцэс (12-р сар)', lambda d: d['is_year_end'].sum()),
                        ('desc_mismatch', '⚠️ Тайлбар ↔ дансны хэв маяг зөрсөн', lambda d: d['desc_mismatch'].sum()),
                        ('name_no_overlap', '⚠️ Дансны нэр ↔ тайлбар давхцахгүй', lambda d: d['name_no_overlap'].sum()),
                        ('dir_mismatch', '⚠️ Дансны төрөл ↔ чиглэл зөрсөн', lambda d: d['dir_mismatch'].sum()),
                    ]
                    for code, label, fn in feat_info:
                        if code in txn_result.columns:
                            cnt = fn(txn_result)
                            pct = cnt / len(txn_result) * 100
                            feat_summary.append({'Шинж чанар': label, 'Илэрсэн тоо': f"{cnt:,}", 'Хувь': f"{pct:.1f}%"})
                    st.dataframe(pd.DataFrame(feat_summary), use_container_width=True, hide_index=True)

                # Жагсаалт
                st.markdown("---")
                txn_years = sorted(txn_result['report_year'].dropna().unique().tolist()) if 'report_year' in txn_result.columns else []
                fc1, fc2 = st.columns(2)
                with fc1:
                    risk_f = st.selectbox("Эрсдэлийн түвшин:", ['Бүгд','🔴 Маш өндөр','🟠 Өндөр','🟡 Дунд'], key='txn_risk_f')
                with fc2:
                    year_f = st.selectbox("Он:", ['Бүгд'] + [str(y) for y in txn_years], key='txn_year_f')
                t_show = txn_result[txn_result['txn_anomaly']==1].copy() if risk_f=='Бүгд' else txn_result[txn_result['txn_risk_level']==risk_f].copy()
                if year_f != 'Бүгд' and 'report_year' in t_show.columns:
                    t_show = t_show[t_show['report_year'].astype(str)==year_f]
                cols_show = ['txn_risk_level','txn_risk','report_year','account_code','account_name','counterparty_name','transaction_date','debit_mnt','credit_mnt','transaction_description','desc_mismatch','name_no_overlap','dir_mismatch','amt_zscore','is_dup','cp_rare']
                t_disp = t_show[[c for c in cols_show if c in t_show.columns]].sort_values('txn_risk', ascending=False)
                st.write(f"Нийт: **{len(t_disp):,}** гүйлгээ")
                st.dataframe(t_disp, use_container_width=True, hide_index=True, height=500)
                st.download_button("📥 Хэвийн бус гүйлгээ CSV", t_disp.to_csv(index=False).encode('utf-8-sig'), "anomaly_txn.csv")
            next_idx += 1

            with all_tabs[next_idx]:
                st.subheader("👤 Харилцагчаар нэгтгэсэн эрсдэл")
                txn_years2 = sorted(txn_result['report_year'].dropna().unique().tolist()) if 'report_year' in txn_result.columns else []
                year_f2 = st.selectbox("Он:", ['Бүгд'] + [str(y) for y in txn_years2], key='cp_year_f')
                txn_filtered = txn_result.copy()
                if year_f2 != 'Бүгд' and 'report_year' in txn_filtered.columns:
                    txn_filtered = txn_filtered[txn_filtered['report_year'].astype(str)==year_f2]
                cp_r = txn_filtered[txn_filtered['counterparty_name'].fillna('')!=''].groupby('counterparty_name').agg(
                    total=('amount','count'), anomaly=('txn_anomaly','sum'), amount=('amount','sum'),
                    accounts=('account_code','nunique'), desc_mis=('desc_mismatch','sum'), dir_mis=('dir_mismatch','sum')
                ).reset_index()
                cp_r['anomaly_pct'] = (cp_r['anomaly']/cp_r['total']*100).round(1)
                cp_r = cp_r.sort_values('anomaly', ascending=False)
                cp_r.columns = ['Харилцагч','Нийт гүйлгээ','Хэвийн бус','Нийт дүн','Дансны тоо','Тайлбар зөрчил','Чиглэл зөрчил','Хэвийн бус %']
                st.write(f"Нийт: **{len(cp_r):,}** харилцагч")
                st.dataframe(cp_r.head(50), use_container_width=True, hide_index=True)
                st.download_button("📥 Харилцагчийн жагсаалт CSV", cp_r.to_csv(index=False).encode('utf-8-sig'), "counterparty_risk.csv", key='dl_cp')
                top20 = cp_r.head(20)
                if len(top20) > 0:
                    st.plotly_chart(px.bar(top20, x='Хэвийн бус', y='Харилцагч', orientation='h', color='Дансны тоо', color_continuous_scale='Reds', title='Топ 20 — хэвийн бус гүйлгээтэй харилцагч').update_layout(height=500, yaxis={'categoryorder':'total ascending'}), use_container_width=True)
                # Дансаар
                st.markdown("---")
                st.subheader("🏷️ Дансаар нэгтгэсэн эрсдэл")
                acct_r = txn_filtered.groupby(['account_code','account_name']).agg(
                    total=('amount','count'), anomaly=('txn_anomaly','sum'), desc_mis=('desc_mismatch','sum'), dir_mis=('dir_mismatch','sum')
                ).reset_index()
                acct_r['anomaly_pct'] = (acct_r['anomaly']/acct_r['total']*100).round(1)
                acct_r = acct_r.sort_values('anomaly', ascending=False)
                acct_r.columns = ['Дансны код','Дансны нэр','Нийт','Хэвийн бус','Тайлбар зөрчил','Чиглэл зөрчил','Хэвийн бус %']
                st.write(f"Нийт: **{len(acct_r):,}** данс")
                st.dataframe(acct_r.head(50), use_container_width=True, hide_index=True)
                st.download_button("📥 Дансны жагсаалт CSV", acct_r.to_csv(index=False).encode('utf-8-sig'), "account_risk.csv", key='dl_acct')
            next_idx += 1

        if has_rm:
            with all_tabs[next_idx]:
                td.show_risk_matrix_description()
                st.subheader("🎯 Эрсдэлийн матриц")
                rm_all['risk_score'] = pd.to_numeric(rm_all['risk_score'], errors='coerce').fillna(0)
                rm_all['total_amount_mnt'] = pd.to_numeric(rm_all.get('total_amount_mnt', 0), errors='coerce').fillna(0)
                rm_summary = []
                for yv in sorted(rm_all['year'].unique()):
                    rmy = rm_all[rm_all['year'] == yv]
                    rm_summary.append({'Жил': yv, 'Нийт хос': f"{len(rmy):,}", 'Эрсдэлтэй': f"{len(rmy[rmy['risk_score']>0]):,}", 'Хувь': f"{len(rmy[rmy['risk_score']>0])/max(len(rmy),1)*100:.1f}%"})
                st.dataframe(pd.DataFrame(rm_summary), use_container_width=True, hide_index=True)
                fig_rm = go.Figure()
                for yv in sorted(rm_all['year'].unique()):
                    rmy = rm_all[rm_all['year'] == yv]
                    fig_rm.add_trace(go.Bar(x=['Нийт', 'Эрсдэлтэй'], y=[len(rmy), len(rmy[rmy['risk_score'] > 0])], name=str(yv)))
                fig_rm.update_layout(barmode='group', height=350)
                st.plotly_chart(fig_rm, use_container_width=True)
                st.subheader("Топ 20 харилцагч")
                top_cp = rm_all.groupby('counterparty_name').agg(txn=('transaction_count', 'sum'), accounts=('account_code', 'nunique')).sort_values('txn', ascending=False).head(20).reset_index()
                top_cp.columns = ['Харилцагч', 'Гүйлгээний тоо', 'Дансны тоо']
                st.dataframe(top_cp, use_container_width=True, hide_index=True)
                td.show_risk_matrix_interpretation(n_pairs=len(rm_all))
            next_idx += 1

        if has_mo:
            with all_tabs[next_idx]:
                td.show_monthly_trend_description()
                st.subheader("📈 Сарын чиг хандлага")
                mo_all['total_debit_mnt'] = pd.to_numeric(mo_all['total_debit_mnt'], errors='coerce').fillna(0)
                mo_all['transaction_count'] = pd.to_numeric(mo_all['transaction_count'], errors='coerce').fillna(0)
                mo_agg = mo_all.groupby('month').agg(debit=('total_debit_mnt', 'sum'), txn=('transaction_count', 'sum')).reset_index()
                mo_agg['debit_T'] = mo_agg['debit'] / 1e9
                fig_mo = make_subplots(rows=2, cols=1, subplot_titles=("Эргэлт (T₮)", "Гүйлгээний тоо"))
                fig_mo.add_trace(go.Scatter(x=mo_agg['month'], y=mo_agg['debit_T'], name='Дебит'), row=1, col=1)
                fig_mo.add_trace(go.Bar(x=mo_agg['month'], y=mo_agg['txn'], name='Гүйлгээ'), row=2, col=1)
                fig_mo.update_layout(height=500)
                st.plotly_chart(fig_mo, use_container_width=True)
                td.show_monthly_trend_interpretation()

        td.show_dashboard_footer()

    if not st.session_state.get('analysis_done', False) and not has_any:
        st.info("👆 Файлуудаа оруулна уу. TB + Ledger = бүрэн шинжилгээ. ЕДТ = гүйлгээний шинжилгээ.")
